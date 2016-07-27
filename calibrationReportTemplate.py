#! /usr/bin/env python3
# calibrationReportTemplate.py

"""
THIS IS A PRACTICE FILE - NOT FINAL!!!
"""
Assumptions:
The rows are always the same:
*blank row
*header row
*row for each of 23 behaviors
*summary rows
--Total Ratings (formula)
--Total Matches (code)
--Total Variances (formula)
--Match % (formula)
--Variance % (formula)

"""

# TODO - Copy ratings from "standard" report (report with correct ratings) to report
# TODO - Copy ratings from each submitted input workbook to report
# Will need to check for next available column so as not to overwrite previous one
# TODO - Check for a next available row (if submissions > # rows avail, warning or copy column
# Input or check field in which you can put number of submissions?

# TODO - Hide unused rows

# TODO - Calculate number of variances

# TODO - Nice to have - highlight individual variances from standard rating


# TODO - set up arguments to capture workbook name and tab name(s)
# TODO - hardcoded workbook name and tab name below for testing purposes

import openpyxl
import os
# Open input file, copy/save as output file
# Overwrite variables so file used in rest of script is copy (i.e., output file)
filename = 'calibrationReportTemplateInputTest.xlsx'
fullFilename = os.path.abspath(filename)
wb = openpyxl.load_workbook(fullFilename)
filename = 'calibrationReportTemplateOutputTest.xlsx'
wb.save(filename)
fullFilename = os.path.abspath(filename)
wb = openpyxl.load_workbook(fullFilename)

# Cycle through each tab with call results and take actions 
# TODO - Build tabs based on pre-defined number of calls or find tabs with data populated; obtain tab name automatically instead of hard-coding as did below
tabName = 'Call 1 Results'
sheet = wb.get_sheet_by_name(tabName)

# Create list for behaviors
# Behaviors are in Column A, rows 3 - 25
behaviors = []
for rowNumber in range(3, 26):
    behaviors.append(sheet.cell(row=rowNumber, column=1).value)
    
#print(behaviors)
#print(len(behaviors))

"""
Commenting out for now; later code that creates dictionary for each submittor can also be used for dict for standards

standardRatings = []
for rowNumber in range(3, 26):
    standardRatings.append(sheet.cell(row=rowNumber, column=7).value)

#print(len(standardRatings))
#print(standardRatings)
if len(behaviors) == len(standardRatings):
    zippedRatings = zip(behaviors, standardRatings)
# from two lists, zip creates a list of paired tuples
# we can then use the pairings to create a dict
standardRatingsDict = dict(zippedRatings)
"""

# Obtain value of non-blank rating row: print(len(sheet.cell(row=2, column=8).value))
# Obtain value of blank rating row: Value is 7: print(len(sheet.cell(row=2, column=11).value))

# Create list of submittor names with which to name dictionary
# Create the dict for each submission with ratings
submissionNames = []
for column in range(7, 26):
    submittorName = sheet.cell(row=2, column=column).value
    if submittorName != 'Rating-':
        submissionNames.append(submittorName)

# Some submission names have spaces; replace space with underscore
firstCleanedUpSubmissionNames = []
for submittorName in submissionNames:
    interimTuple = submittorName.split()
    firstCleanedUpSubmissionNames.append('_'.join(interimTuple))
    
finalCleanedUpSubmissionNames = []
for name in firstCleanedUpSubmissionNames:
    interimTuple = name.split('-')
    finalCleanedUpSubmissionNames.append('_'.join(interimTuple))

#print(finalCleanedUpSubmissionNames)

# Create a blank dictionary with the dict name as the submittor name

submissions = {}
for name in finalCleanedUpSubmissionNames:
    submissions[name] = {}

"""
# Iterate each key in submissions, i.e., each submittor name
#for key in submissions.keys():
for name in finalCleanedUpSubmissionNames:
# For each row of behaviors
    for row in range(3, 26):
# For each submittor column
        for column in range(7, 28):
# Check if column heading, i.e., submittor, matches the key
            if sheet.cell(row=2, column=column).value == name:
# Add the behavior:rating pair to that key's value dictionary
# Behaviors are in first column (1); ratings are in columns ranging from 7 to 28
                submissions[name][sheet.cell(row=row, column=1).value] =  sheet.cell(row=row, column=column).value

print(submissions)
"""
# TEST

"""
print(finalCleanedUpSubmissionNames)

for column in range(7, 28):
    print(str(sheet.cell(row=2, column=column).value))
print(submissions)
"""

"""
for name in finalCleanedUpSubmissionNames:
    #for column in range(7,8):
    print(name == sheet.cell(row=2, column=7).value)
"""
"""prints: 
False
False
False
False
"""
"""
for name in submissions.keys():
    print(name == sheet.cell(row=2, column=7).value)    
"""
"""
prints:

False
False
False
False
"""
"""
    Stop @ 08:45pm - printed:
{'Rating_Miguel_Terrazas': {}, 'Rating_Standard': {}, 'Rating_Doriana_Ciuciu': {}, 'Rating_Marius_Dinu': {}}


print(submissions.keys()):
dict_keys(['Rating_Miguel_Terrazas', 'Rating_Marius_Dinu', 'Rating_Standard', 'Rating_Doriana_Ciuciu'])
"""
"""
Problem: Can't at runtime create variable name from a string and use it as the dict name.
Solution: Create a dictionary of dictionaries.
Action: Above, create code to create dict of dicts. Below, clean up code to work within that framework.
Structure: The key is the  submittor name (e.g., 'Rating_Marius_Dinu'). The value for that key is the dictionary of 'behavior:rating' pairs.

for name in finalCleanedUpSubmissionNames:
    name = {}

# Add "behavior:rating" key:value pair to the dictionary


for name in finalCleanedUpSubmissionNames:
    for column in range(8, 26):
        if sheet.cell(row=2, column=column).value == name:
            for row in range(3, 26):
                name[sheet.cell(row=row, column=1)] = sheet.cell(row=row, column=column).value 

"""
# Use list in submission names to find out how many columns to review
# Start at column 8 (first submission name)
"""
variances_list = []
for column in range(8, len(submissionNames)):
    variances = 0
    for row in range(3, 26):
        if sheet.cell(row=row, column=column).value != sheet.cell(row=row, column=7):
            variances += 1
    #sheet.cell(column=column, row=28, value="%s" % variances)
    variances_list.append(variances)
"""
#print(variances_list)    
        
# TEST
#print(submissionNames)
#print(len(submissionNames))

#print(sheet.cell(row=3, column=8).value == sheet.cell(row=3, column=7).value)
variances_list = []
# Note for testing only - submissionNames includes Rating-Standard
# For new approach (brunt comparision of cell to cell instead of building dictionaries, don't need Rating-Standard)
# So subtracting one from length of submission names instead of removing Rating-Standard (in case want to reuse with standard later)
for column in range(8, 8+(len(submissionNames))-1):
    variances = 0
    for row in range(3, 26):
        if sheet.cell(row=row, column=column).value != sheet.cell(row=row, column=7).value:
            variances += 1
    variances_list.append(variances)

variancesDict = {}
for i in range(1, len(submissionNames)):
    variancesDict[submissionNames[i]] = variances_list[i-1]


for column in range(8, 8+(len(variancesDict))):
    for k, v in variancesDict.items():
        if k == sheet.cell(row=2, column=column).value:
            sheet.cell(row=28, column=column).value = v

#TEST populating cell with value (since above failed)
#sheet.cell(row=28, column=11).value = 2
wb.save(fullFilename)
"""
TEST:
print(variancesDict)

for k, v in variancesDict.items():
    print(k, v)

for k, v in variancesDict.items():
    if k == sheet.cell(row=2, column=8).value:
        print('Found match in spreadsheet:%s, %s' % (k, v))


# Testing of value class for spreadsheet cell objects
# Testing of comparison of cell values to values in dictionaries and lists

print(variancesDict)
print(submissionNames)
print(type(sheet.cell(row=2, column=8).value))
print(sheet.cell(row=2, column=8).value in variancesDict)
print(sheet.cell(row=2, column=8).value in submissionNames)

Output: 
'Rating-Marius Dinu': 2, 'Rating-Doriana Ciuciu': 2, 'Rating-Miguel Terrazas': 2}
['Rating-Standard', 'Rating-Marius Dinu', 'Rating-Doriana Ciuciu', 'Rating-Miguel Terrazas']
<class 'str'>
True
True
"""

